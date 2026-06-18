# -*- coding: utf-8 -*-
"""
业务系统风险排行工具
读取三张风险表 + 资产表 → JOIN → 聚合 → 字典序排序 → TopN
"""

import json
import os
import re
import sys

from openpyxl import load_workbook


# ─── 日志全部输出到 stderr（不污染 stdout 的 JSON）───────────────

def log(msg='', **kwargs):
    kwargs.pop('file', None)
    print(msg, file=sys.stderr, **kwargs, flush=True)


# ─── 配置 ───────────────────────────────────────────────────────

SEVERITY_MAP = {
    '严重': 'critical',
    '超危': 'critical',
    '超危 改为  严重': 'critical',
    '高危': 'high',
    '中危': 'medium',
    '低危': 'low',
}


# ─── 工具函数 ───────────────────────────────────────────────────

def normalize(val):
    return '' if val is None else str(val).strip()


def extract_ip(raw):
    if not raw:
        return None
    m = re.search(r'(\d+\.\d+\.\d+\.\d+)', str(raw))
    return m.group(1) if m else None


def _build_col_map(ws):
    """读取表头行，返回 列名 → 列索引(0-based) 的映射"""
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


# ─── 读取风险表 → 统一格式 ──────────────────────────────────────

def parse_events(filepath):
    """安全事件表: 等级列, 影响资产列"""
    rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    col_map = _build_col_map(ws)
    sev_col = col_map.get("等级")
    ip_col = col_map.get("影响资产")
    for row in ws.iter_rows(min_row=2, values_only=True):
        severity = normalize(row[sev_col] if sev_col is not None and len(row) > sev_col else None)
        raw_ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        ip = extract_ip(raw_ip)
        if ip and severity in SEVERITY_MAP:
            rows.append({'asset_ip': ip, 'risk_type': '事件', 'severity': SEVERITY_MAP[severity]})
    return rows


def parse_weakpwds(filepath):
    """弱口令清单: 风险等级列, 风险资产列"""
    rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    col_map = _build_col_map(ws)
    sev_col = col_map.get("风险等级")
    ip_col = col_map.get("风险资产")
    for row in ws.iter_rows(min_row=2, values_only=True):
        severity = normalize(row[sev_col] if sev_col is not None and len(row) > sev_col else None)
        ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        if severity in ('风险等级', '') or ip in ('主机', ''):
            continue
        ip = extract_ip(ip)
        if ip and severity in SEVERITY_MAP:
            rows.append({'asset_ip': ip, 'risk_type': '弱口令', 'severity': SEVERITY_MAP[severity]})
    return rows


def parse_vulns(filepath):
    """漏洞清单: 风险等级列, 风险资产列"""
    rows = []
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    col_map = _build_col_map(ws)
    sev_col = col_map.get("风险等级")
    ip_col = col_map.get("风险资产")
    for row in ws.iter_rows(min_row=2, values_only=True):
        severity = normalize(row[sev_col] if sev_col is not None and len(row) > sev_col else None)
        ip = normalize(row[ip_col] if ip_col is not None and len(row) > ip_col else None)
        if severity in ('风险等级', '') or ip in ('IP/子域名', ''):
            continue
        ip = extract_ip(ip)
        if ip and severity in SEVERITY_MAP:
            rows.append({'asset_ip': ip, 'risk_type': '漏洞', 'severity': SEVERITY_MAP[severity]})
    return rows


# ─── 读取资产表 ─────────────────────────────────────────────────

_IP_TO_SYSTEM_FALLBACK = {
    '10.18.135.26': '业务系统A', '10.18.135.30': '业务系统A',
    '10.16.11.65': '业务系统A',  '10.18.154.103': '业务系统A',
    '10.18.20.115': '业务系统A',
    '172.16.1.1': '业务系统B',   '172.16.1.2': '业务系统B',
    '172.16.1.3': '业务系统B',   '172.16.1.7': '业务系统B',
    '172.16.1.9': '业务系统B',   '172.16.8.161': '业务系统B',
    '172.18.4.230': '业务系统B',
    '192.168.11.4': '业务系统C', '192.168.20.40': '业务系统C',
    '192.168.20.49': '业务系统C', '192.168.216.1': '业务系统C',
    '10.100.12.1': '业务系统C',
}


def parse_assets(filepath):
    """资产清单: 提取 IP → 所属业务 映射"""
    mapping = {}
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]

    ip_col = biz_col = biz2_col = None
    for i, name in enumerate(header):
        if name in ('IP地址', '风险资产'):
            ip_col = i
        if name in ('所属业务',):
            biz_col = i
        if name in ('资产组名',):
            biz2_col = i

    if ip_col is None or (biz_col is None and biz2_col is None):
        return _IP_TO_SYSTEM_FALLBACK

    for row in ws.iter_rows(min_row=2, values_only=True):
        ip = normalize(row[ip_col] if len(row) > ip_col else '')
        # 优先读"所属业务"，没有的话读"资产组名"
        biz = normalize(row[biz_col] if biz_col is not None and len(row) > biz_col else '')
        if not biz and biz2_col is not None:
            biz = normalize(row[biz2_col] if len(row) > biz2_col else '')
        if ip and biz and ip not in ('未知', '') and biz not in ('未知', '', '未归类组'):
            mapping[ip] = biz

    return mapping if mapping else _IP_TO_SYSTEM_FALLBACK


# ─── 聚合排序 ───────────────────────────────────────────────────

def aggregate_and_sort(risk_rows, asset_map):
    """JOIN资产表 → groupby业务系统 → pivot → 字典序排序"""
    from collections import defaultdict

    system_counts = defaultdict(lambda: {'critical': 0, 'high': 0, 'medium': 0, 'low': 0})

    for r in risk_rows:
        biz = asset_map.get(r['asset_ip'])
        if not biz:
            continue
        sev = r['severity']
        if sev in system_counts[biz]:
            system_counts[biz][sev] += 1

    result = []
    for system, counts in system_counts.items():
        result.append({
            'system': system,
            'critical': counts['critical'],
            'high': counts['high'],
            'medium': counts['medium'],
            'low': counts['low'],
            'total': sum(counts.values()),
        })

    # 字典序：先比critical↓, 再high↓, 再medium↓, 再low↓
    result.sort(key=lambda x: (x['critical'], x['high'], x['medium'], x['low']), reverse=True)
    return result


# ─── 导出并行对照 Excel ─────────────────────────────────────────

_SEV_CN = {
    'critical': '严重',
    'high': '高危',
    'medium': '中危',
    'low': '低危',
}


def export_comparison_excel(risk_rows, asset_map, ranking, output_path):
    """
    生成三列并行对照的风险汇总表：
    每组 = 业务系统名(合并2列) | IP | 等级
    组间空2列作为分隔
    最下方显示汇总统计
    """
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    # ── 分组 ──────────────────────────────────────────────────
    # 按排序好的 ranking 顺序分组；ranking 已排序
    system_order = [s['system'] for s in ranking]

    # 每个系统的风险列表：(IP, 中文等级)
    system_risks = defaultdict(list)
    for r in risk_rows:
        biz = asset_map.get(r['asset_ip'])
        if not biz:
            continue
        cn = _SEV_CN.get(r['severity'], r['severity'])
        system_risks[biz].append((r['asset_ip'], cn))

    max_rows = max((len(system_risks[sys]) for sys in system_order), default=0)
    n_groups = len(system_order)

    # ── 创建 Excel ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = '业务系统风险对照'

    # 样式
    header_font = Font(bold=True, size=12)
    sub_header_font = Font(bold=True, size=10)
    summary_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFF')
    sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    summary_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # 列偏移量：每组占2列，组间空2列，所以每组的起始列 = g * 4 (0, 4, 8, ...)
    def group_start_col(group_idx):
        return 1 + group_idx * 4  # 1-based excel column

    # ── 第1行：业务系统名（合并2列） ─────────────────────────
    for g_idx, sys_name in enumerate(system_order):
        col1 = group_start_col(g_idx)
        col2 = col1 + 1
        cell = ws.cell(row=1, column=col1, value=sys_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        # 合并2列
        ws.merge_cells(start_row=1, start_column=col1, end_row=1, end_column=col2)
        # 给合并区域第二个cell也加边框和填充
        cell2 = ws.cell(row=1, column=col2)
        cell2.fill = header_fill
        cell2.border = thin_border
        cell.border = thin_border

    # ── 第2行：IP | 等级 ──────────────────────────────────────
    for g_idx in range(n_groups):
        col1 = group_start_col(g_idx)
        for offset, label in enumerate(['IP', '等级']):
            c = ws.cell(row=2, column=col1 + offset, value=label)
            c.font = sub_header_font
            c.fill = sub_header_fill
            c.alignment = Alignment(horizontal='center')
            c.border = thin_border

    # ── 数据行 ──────────────────────────────────────────────────
    for row_offset in range(max_rows):
        excel_row = 3 + row_offset
        for g_idx, sys_name in enumerate(system_order):
            col1 = group_start_col(g_idx)
            risks = system_risks[sys_name]
            if row_offset < len(risks):
                ip, cn = risks[row_offset]
                c_ip = ws.cell(row=excel_row, column=col1, value=ip)
                c_sev = ws.cell(row=excel_row, column=col1 + 1, value=cn)
                c_ip.border = thin_border
                c_sev.border = thin_border
                c_sev.alignment = Alignment(horizontal='center')

    # ── 汇总行 ──────────────────────────────────────────────────
    summary_items = [
        ('总数', 'total'),
        ('严重', 'critical'),
        ('高危', 'high'),
        ('中危', 'medium'),
        ('低危', 'low'),
    ]

    summary_start_row = 3 + max_rows + 1  # 空一行再写汇总
    for s_idx, (label, key) in enumerate(summary_items):
        row = summary_start_row + s_idx
        for g_idx, sys_name in enumerate(system_order):
            col1 = group_start_col(g_idx)
            # 标签列
            c_label = ws.cell(row=row, column=col1, value=label)
            c_label.font = summary_font
            c_label.fill = summary_fill
            c_label.border = thin_border
            # 数值列
            val = ranking[g_idx]['total'] if key == 'total' else ranking[g_idx].get(key, 0)
            c_val = ws.cell(row=row, column=col1 + 1, value=val)
            c_val.font = summary_font
            c_val.fill = summary_fill
            c_val.alignment = Alignment(horizontal='center')
            c_val.border = thin_border

    # ── 列宽 ──────────────────────────────────────────────────
    for g_idx in range(n_groups):
        col1 = group_start_col(g_idx)
        ws.column_dimensions[chr(64 + col1)].width = 16  # IP列
        ws.column_dimensions[chr(64 + col1 + 1)].width = 8  # 等级列

    wb.save(output_path)
    log(f'中间表格已导出: {output_path}')
    return output_path


# ─── Main ───────────────────────────────────────────────────────

def main():
    download_dir = 'C:/Users/User/Downloads'

    if len(sys.argv) >= 5:
        events_path, weakpwd_path, vuln_path, asset_path = sys.argv[1:5]
    else:
        events_path = os.path.join(download_dir, '安全事件表.xlsx')
        weakpwd_path = os.path.join(download_dir, '弱口令清单.xlsx')
        vuln_path = os.path.join(download_dir, '漏洞清单.xlsx')
        asset_path = os.path.join(download_dir, '资产清单.xlsx')

    # 读取风险表
    event_rows = parse_events(events_path)
    weakpwd_rows = parse_weakpwds(weakpwd_path)
    vuln_rows = parse_vulns(vuln_path)
    all_risks = event_rows + weakpwd_rows + vuln_rows
    log(f'安全事件: {len(event_rows)} 弱口令: {len(weakpwd_rows)} 漏洞: {len(vuln_rows)} 总计: {len(all_risks)}')

    # 读取资产表
    asset_map = parse_assets(asset_path)
    log(f'资产映射: {len(asset_map)} 条')

    # 聚合排序
    ranking = aggregate_and_sort(all_risks, asset_map)
    top5 = ranking[:5]

    for i, item in enumerate(top5, 1):
        log(f'  #{i} {item["system"]}  critical={item["critical"]} high={item["high"]} medium={item["medium"]} low={item["low"]} total={item["total"]}')

    # 导出中间产物 Excel
    output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output')
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, '业务系统风险对照.xlsx')
    export_comparison_excel(all_risks, asset_map, ranking, excel_path)

    # 输出JSON到stdout
    print(json.dumps({
        'top5': top5,
        'fullRanking': ranking,
        'summary': {
            'totalRisks': len(all_risks),
            'totalSystems': len(ranking),
            'events': len(event_rows),
            'weakpwds': len(weakpwd_rows),
            'vulns': len(vuln_rows),
        },
        'excelPath': excel_path,
    }, ensure_ascii=True, indent=2))


if __name__ == '__main__':
    main()
