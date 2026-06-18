import json
import sys

from openpyxl import load_workbook


def normalize(value):
    return "" if value is None else str(value).strip()


def parse_number(value):
    text = normalize(value)
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def _build_col_map(ws):
    """读取表头行，返回 列名 → 列索引(0-based) 的映射"""
    header = [normalize(cell) for cell in next(ws.iter_rows(values_only=True))]
    return {name: i for i, name in enumerate(header) if name}


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: incident_status_stats.py <incident.xlsx>")

    workbook = load_workbook(sys.argv[1], read_only=True, data_only=True)
    sheet = workbook.active

    col_map = _build_col_map(sheet)
    level_col = col_map.get("等级")
    status_col = col_map.get("处置状态")
    contain_col = col_map.get("遏制时长(秒)")
    ip_col = col_map.get("影响资产")

    total = 0
    severe = 0
    high = 0
    closed = 0
    processing = 0
    unique_ips = set()
    contain_total = 0.0
    contain_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(normalize(cell) for cell in row):
            continue

        total += 1
        if level_col is not None and len(row) > level_col:
            level = normalize(row[level_col])
            if level == "严重":
                severe += 1
            elif level == "高危":
                high += 1

        if status_col is not None and len(row) > status_col:
            status = normalize(row[status_col])
            if status == "已闭环":
                closed += 1
            elif status == "处置中":
                processing += 1

        if contain_col is not None and len(row) > contain_col:
            contain_min = parse_number(row[contain_col])
            if contain_min is not None:
                contain_total += contain_min
                contain_count += 1

        if ip_col is not None and len(row) > ip_col:
            raw_ip = normalize(row[ip_col])
            if raw_ip:
                ip_only = raw_ip.split("(")[0].strip()
                if ip_only:
                    unique_ips.add(ip_only)

    close_rate = round((closed / total) * 100) if total else 0
    average_contain_min = round(contain_total / contain_count) if contain_count else 0
    print(json.dumps({
        "totalEvents": total,
        "severeEvents": severe,
        "highEvents": high,
        "closedEvents": closed,
        "processingEvents": processing,
        "closeRate": close_rate,
        "uniqueAssetCount": len(unique_ips),
        "averageContainMin": average_contain_min
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
